"""
src/case_tables.py
==================

Where a case's three tables live, and how to read them.

WHY A WORKBOOK
--------------
`source`, `processes` and `TCs` describe one case and are edited together, by a
person, by hand. Three separate CSVs make that harder than it needs to be: no
dropdown to stop a flow name being mistyped, no room for a note beside a
number, and three files to keep consistent with each other.

So a case may carry a single `input_data/case.xlsx` with one sheet per table.
The workbook is the input -- there is no export step and no second copy, because
two copies of one table, one generated from the other, is exactly the drift this
project keeps finding.

CSV IS STILL READ
-----------------
A case with `source.csv`, `processes.csv` and `TCs.csv` works unchanged. The
reference fixtures under `data_folder/reference/` are CSV and stay that way:
every test suite runs on them, and converting the thing that proves the code
works would be a poor trade for tidiness.

A case may not have both for the same table. Two files of the same name with
different contents is precisely the situation where someone edits one and the
model reads the other.

EMPTY CELLS
-----------
A CSV read with `keep_default_na=False` gives `''` for a blank. Excel gives
`NaN`. The model reads `''` as "this layer is not populated" -- the whole
nesting rule depends on it -- so blanks are normalised here rather than in
every caller.
"""
from __future__ import annotations

import os

import pandas as pd

WORKBOOK = 'case.xlsx'

# Sheet name and CSV name are the same word, so a case reads the same either way.
TABLES = ('source', 'processes', 'TCs')

CSV_OPTIONS = dict(keep_default_na=False, na_values=[])


class CaseTableError(ValueError):
    """Raised when a case's tables cannot be located, or are offered twice."""


def workbook_path(case: str) -> str:
    return os.path.join(case, 'input_data', WORKBOOK)


def csv_path(case: str, table: str) -> str:
    return os.path.join(case, 'input_data', f'{table}.csv')


def has_workbook(case: str) -> bool:
    return os.path.exists(workbook_path(case))


def _sheet_names(path: str) -> list[str]:
    from openpyxl import load_workbook
    book = load_workbook(path, read_only=True)
    try:
        return list(book.sheetnames)
    finally:
        book.close()


def where(case: str, table: str) -> tuple[str, str] | None:
    """
    ('xlsx', path) or ('csv', path) for one table, or None if it has neither.

    Raises when both exist: that is not a preference to resolve silently.
    """
    if table not in TABLES:
        raise CaseTableError(f'{table!r} is not one of {", ".join(TABLES)}')

    book, delimited = workbook_path(case), csv_path(case, table)
    in_book = os.path.exists(book) and table in _sheet_names(book)
    in_csv = os.path.exists(delimited)

    if in_book and in_csv:
        raise CaseTableError(
            f"{case} holds {table} twice: sheet '{table}' in {WORKBOOK}, and "
            f"{table}.csv beside it. Keep one -- otherwise the model reads a "
            f"table that may not be the one being edited.")

    if in_book:
        return ('xlsx', book)
    if in_csv:
        return ('csv', delimited)
    return None


def exists(case: str, table: str) -> bool:
    return where(case, table) is not None


def read(case: str, table: str, dtype=None) -> pd.DataFrame:
    """One of a case's tables, from whichever format it is kept in."""
    found = where(case, table)
    if found is None:
        raise CaseTableError(
            f'{case} has no {table}: expected sheet {table!r} in '
            f'{workbook_path(case)}, or {csv_path(case, table)}.')

    kind, path = found
    if kind == 'csv':
        return pd.read_csv(path, dtype=dtype, **CSV_OPTIONS)

    frame = pd.read_excel(path, sheet_name=table, dtype=dtype)
    return normalise(frame)


def _as_text(value) -> str:
    """One cell as the CSV reader would have produced it."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ''
    if isinstance(value, bool):
        return '1' if value else ''
    if isinstance(value, float) and value.is_integer():
        # Excel stores every number as a float, so a column of 1s arrives as
        # 1.0. Written back out as '1.0' it stops matching the '1' that
        # is_residual is tested against, and residual rows go unrecognised.
        return str(int(value))
    return str(value).strip()


def normalise(frame: pd.DataFrame) -> pd.DataFrame:
    """
    Make a sheet read like the CSV it replaces.

    The CSV reader runs with `keep_default_na=False`, so a column holding any
    blank comes back as strings -- '' and '1' -- while a fully populated
    numeric column comes back numeric. Excel has no empty string: a blank is
    NaN, which makes the whole column float64, and '1' becomes 1.0.

    That difference is not cosmetic. `is_residual` is tested with
    `str(value).strip() in ('1', 'True', 'true')`; against 1.0 that is '1.0',
    which matches nothing, and the Monte Carlo then refuses a group whose
    residual rows it can no longer identify. So columns containing a blank are
    converted to text here, and fully numeric columns are left numeric --
    exactly what reading the CSV produced.
    """
    frame = frame.copy()
    # A wholly empty trailing row is what Excel leaves behind after a delete.
    frame = frame.dropna(how='all')

    for column in frame.columns:
        has_blank = frame[column].isna().any()
        if frame[column].dtype == object or has_blank:
            frame[column] = frame[column].map(_as_text)

    return frame.reset_index(drop=True)


def describe(case: str) -> str:
    """Which file backs each table, for a run to report."""
    parts = []
    for table in TABLES:
        found = where(case, table)
        parts.append(f'{table}={found[0] if found else "missing"}')
    return ', '.join(parts)


# ----------------------------------------------------------------------
# Writing
# ----------------------------------------------------------------------

# Allowed values live on their own hidden sheet and the dropdowns point at
# ranges on it. An inline list -- formula1='"a,b,c"' -- is capped at 255
# characters, which a real element list passes without warning: Excel then
# drops the validation silently and the dropdown simply is not there.
LISTS_SHEET = '_lists'


def write_sheet(case: str, table: str, frame: pd.DataFrame, *,
                dropdowns: dict[str, list[str]] | None = None,
                widths: dict[str, int] | None = None) -> str:
    """
    Replace one sheet of a case's workbook, leaving the others alone.

    Written to a temp file beside the target and renamed, for the reason in
    tools/make_skeleton.py: a half-written table looks like a smaller table,
    and this one is read back and merged.
    """
    import tempfile

    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    path = workbook_path(case)
    os.makedirs(os.path.dirname(path), exist_ok=True)

    if os.path.exists(path):
        book = load_workbook(path)
        position = book.sheetnames.index(table) if table in book.sheetnames else None
        if position is not None:
            del book[table]
        sheet = book.create_sheet(table, index=position)
    else:
        book = Workbook()
        book.remove(book.active)
        sheet = book.create_sheet(table)

    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False):
        sheet.append(['' if value is None else value for value in row])

    sheet.freeze_panes = 'A2'
    for index, column in enumerate(frame.columns, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = (widths or {}).get(column, 18)

    if dropdowns:
        lists = book[LISTS_SHEET] if LISTS_SHEET in book.sheetnames \
            else book.create_sheet(LISTS_SHEET)
        lists.sheet_state = 'hidden'
        start = lists.max_column + 1 if lists.max_row > 1 else 1

        for offset, (column, allowed) in enumerate(sorted(dropdowns.items())):
            if column not in frame.columns or not allowed:
                continue
            letter = get_column_letter(start + offset)
            lists.cell(row=1, column=start + offset, value=column)
            for line, value in enumerate(allowed, start=2):
                lists.cell(row=line, column=start + offset, value=value)

            rule = DataValidation(
                type='list',
                formula1=f"={LISTS_SHEET}!${letter}$2:${letter}${len(allowed) + 1}",
                allow_blank=True, showDropDown=False)
            # showDropDown=False is Excel's spelling for "do show the arrow";
            # setting it True hides the control while still enforcing the list.
            rule.error = 'Not one of the values this case declares.'
            rule.errorTitle = 'Unknown value'
            sheet.add_data_validation(rule)

            target = get_column_letter(list(frame.columns).index(column) + 1)
            rule.add(f'{target}2:{target}{max(len(frame) + 1, 2)}')

    handle = tempfile.NamedTemporaryFile(
        dir=os.path.dirname(path), prefix='.case-', suffix='.tmp', delete=False)
    handle.close()
    try:
        book.save(handle.name)
        os.replace(handle.name, path)
    except BaseException:
        try:
            os.unlink(handle.name)
        except OSError:
            pass
        raise
    return path
