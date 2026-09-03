"""
Prove the model is not specific to BEV electronics.

    ./.venv/bin/python tests/test_generality.py

WHY THIS EXISTS
---------------
Every other suite runs on vehicle data. They would all keep passing if the code
quietly assumed vehicles, electronics domains, or element names -- which is
exactly the assumption that has to be false for this model to be used on
anything else.

So this one builds a **completely different recovery item** from scratch --
photovoltaic panels, with different groups, different materials and different
elements, in a different unit -- writes it in the upstream format, and puts it
through the same code end to end. Nothing here shares a single name with the
vehicle case.

If it passes, the pipeline is generic. If it fails, the failure names the line
that is still hard-wired.

WHAT IS CHECKED
---------------
1. The upstream reader reads a non-vehicle dataset.
2. The model solves it.
3. Mass balance closes on every year.
4. The Monte Carlo runs on it and produces a spread.
5. `rest` is derived for an incomplete parent, as it is for vehicles.
6. The same fixture read with `child_layer = material` puts the children one
   layer up, with no placeholder -- the 04_01 shape rather than the 04_02 one.
7. Several products in one case each close to 1 on their OWN total, not on the
   pooled one -- the 04_01 drivetrain shape.
8. Draws from two different runs are refused rather than silently averaged
   together, both within one folder and across a case's product folders.

The item is named in the case's own `source.csv`, never in `src/params_schema.py`.
That is what lets 04_01 and 04_02 be different cases rather than different runs
of the same settings, so this suite writes one and would fail if the reader
went back to reading the settings instead.
"""

from __future__ import annotations

import os
import sys

# Run under the project interpreter whatever was typed, and put the repo
# root on the path. Must come before any third-party import.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.bootstrap import ensure_venv
ensure_venv()

import shutil
import tempfile
import traceback

import numpy as np
import pandas as pd

from src.monte_carlo import solve_draws
from src.params_schema import Params
from src.recovery_model_optimized import RecoveryModelOptimized
from src.rest import add_rest

NAMES = ['product', 'component', 'material', 'element']

# A recovery item that is not a vehicle, sharing no name with the real case.
PRODUCT = 'PVPanel'
GROUPS = ('Frame', 'Laminate', 'JunctionBox')
ELEMENTS = {'Frame': {'Al': 1.0},                        # complete: no rest
            'Laminate': {'Si': 0.40, 'Ag': 0.001},       # incomplete -> rest
            'JunctionBox': {'Cu': 0.35, 'Pb': 0.02}}     # incomplete -> rest

# Part of an element's share, resolved into the material holding it -- what
# upstream writes as <element>__<material>__<group>. Shares of the GROUP, so
# each is a slice of that element's entry above: 0.30 of the Laminate is
# silicon in glass feedstock, and the remaining 0.10 of its silicon is not
# resolved into any material. `Frame` has none at all, so it keeps the old
# shape and the two can be compared in one run.
MATERIALS = {'Laminate': {'glassfeed': {'Si': 0.30}},
             'JunctionBox': {'braid': {'Cu': 0.20},
                             'solderbed': {'Cu': 0.05, 'Pb': 0.015}}}
YEARS = (2035, 2045)
DRAWS = 400
FLOW = 'PV_collected'

# Two more products, exported to their own folders the way 04_01 writes one per
# drivetrain. `collected` and `inflow` are the single-product case's folders.
PRODUCTS = ('PVMono', 'PVThin')
FOLDERS = ('collected', 'inflow', *(f'{p}_collected' for p in PRODUCTS))


def write_upstream(root: str, materials: bool = False) -> None:
    """
    A synthetic upstream export, in the layout the reader expects.

    Written with a generator rather than committed as fixtures: the arrays are
    (draws, years) floats and there is no reason to put binaries in git when
    twenty lines reproduce them exactly.

    `materials` adds the `<element>__<material>__<group>` files upstream began
    writing on 2026-08-31. Off by default, so most of this suite keeps reading
    the shape that has no material resolution at all -- which is the shape the
    new one has to reproduce exactly.
    """
    scenario = os.path.join(root, 'HIGH')
    os.makedirs(scenario, exist_ok=True)
    np.save(os.path.join(scenario, 'years.npy'), np.array(YEARS))

    rng = np.random.default_rng(11)
    for flow in FOLDERS:
        folder = os.path.join(scenario, flow)
        os.makedirs(folder, exist_ok=True)
        # Each product at a clearly different level, so reading one folder
        # twice, or pooling the two, shows up as a total that is out by a
        # factor rather than by a rounding error.
        level = 40.0 * 2 ** FOLDERS.index(flow)
        for group in GROUPS:
            # Group mass in kt, growing between the two years, with spread.
            mass = rng.lognormal(mean=np.log([level, level * 2.25]), sigma=0.15,
                                 size=(DRAWS, len(YEARS)))
            np.save(os.path.join(folder, f'__domain____{group}.npy'),
                    mass.astype(np.float32))
            for element, share in ELEMENTS[group].items():
                np.save(os.path.join(folder, f'{element}__{group}.npy'),
                        (mass * share).astype(np.float32))
            if not materials:
                continue
            for material, parts in MATERIALS.get(group, {}).items():
                for element, share in parts.items():
                    np.save(os.path.join(
                        folder, f'{element}__{material}__{group}.npy'),
                        (mass * share).astype(np.float32))


def write_case(case: str, child_layer: str = 'element',
               products: tuple[str, ...] | None = None) -> None:
    """The coefficients and the network, for a panel recycler."""
    os.makedirs(os.path.join(case, 'input_data'), exist_ok=True)

    # What the case is, declared by the case. Deliberately NOT in the settings:
    # a second case must be runnable without editing the first one's numbers.
    pd.DataFrame([
        dict(key='product', value=PRODUCT),
        dict(key='inflow_flow_id', value=FLOW),
        dict(key='flow', value='collected'),
        dict(key='child_layer', value=child_layer),
        dict(key='group_marker', value='__domain__'),
        dict(key='material_suffix', value='_mixed'),
        dict(key='groups', value=''),
    ] if products is None else [
        dict(key='product', value=';'.join(products)),
        dict(key='inflow_flow_id', value=FLOW),
        dict(key='flow', value='{product}_collected'),
        dict(key='child_layer', value=child_layer),
        dict(key='group_marker', value='__domain__'),
        dict(key='material_suffix', value='_mixed'),
        dict(key='groups', value=''),
    ]).to_csv(os.path.join(case, 'input_data', 'source.csv'), index=False)

    # A material-keyed case has no element layer to key on.
    finest = 'element' if child_layer == 'element' else 'material'
    pd.DataFrame([
        dict(Input_FlowID=FLOW, Output_FlowID='PV_delaminated', process='delamination',
             technology='thermal', keyed_at='component', role='intermediate'),
        dict(Input_FlowID=FLOW, Output_FlowID='PV_loss_handling', process='delamination',
             technology='thermal', keyed_at='component', role='loss'),
        dict(Input_FlowID='PV_delaminated', Output_FlowID='PV_recovered',
             process='leaching', technology='hydro', keyed_at=finest,
             role='recovered'),
        dict(Input_FlowID='PV_delaminated', Output_FlowID='PV_loss_leaching',
             process='leaching', technology='hydro', keyed_at=finest,
             role='loss'),
    ]).to_csv(os.path.join(case, 'input_data', 'processes.csv'), index=False)


def settings(root: str, case_name: str) -> Params:
    """A parameter set pointed at the synthetic item, not at the vehicle one."""
    params = Params()
    params.run.data_folder = os.path.join('data_folder', case_name)
    params.run.scenario = 'HIGH'
    params.run.years = ''
    params.run.working_unit = 't'
    params.data.upstream_root = root
    params.data.inflow_draws_dir = '.'
    params.data.upstream_flow = 'collected'
    # NOT set here: product, inflow_flow_id, groups, material_suffix,
    # child_layer. Those come from the case's own source.csv, which is the
    # whole point -- they are left at their vehicle defaults on purpose, so
    # that a reader falling back to the settings names 'BEV' and fails loudly.
    params.data.draws = DRAWS
    return params


def build_everything(child_layer: str = 'element', products=None,
                     materials: bool = False):
    """Set the whole thing up, and return (params, case folder, temp root)."""
    root = tempfile.mkdtemp()
    upstream = os.path.join(root, 'upstream')
    write_upstream(upstream, materials)

    case_name = (f'pv_panels_test_{child_layer}{"_multi" if products else ""}'
                 f'{"_materials" if materials else ""}')
    case = os.path.join('data_folder', case_name)
    if os.path.isdir(case):
        shutil.rmtree(case)
    write_case(case, child_layer, products)

    params = settings(upstream, case_name)
    return params, case, root


def coefficients(case: str, composition: pd.DataFrame) -> None:
    """
    Fill every coefficient the network needs, so the case actually solves.

    Generated from the composition rather than written out, so this stays
    correct if the fixture above gains a group or an element.
    """
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(
        os.path.abspath(__file__))), 'tools'))
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        'make_skeleton', os.path.join(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))), 'tools', 'make_skeleton.py'))
    skeleton = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(skeleton)

    tcs = skeleton.build(case, composition=composition)
    blank = tcs['value'].astype(str).str.strip() == ''
    recovery = tcs['Output_FlowID'].isin(['PV_delaminated', 'PV_recovered'])
    tcs.loc[blank & recovery, ['value', 'value_min', 'value_max']] = ['0.7', '0.5', '0.9']
    tcs.loc[blank & ~recovery, ['value', 'value_min', 'value_max']] = ['0.3', '', '']
    tcs.loc[blank & ~recovery, 'is_residual'] = '1'
    tcs.to_csv(os.path.join(case, 'input_data', 'TCs.csv'), index=False)


# ----------------------------------------------------------------------

def test_a_different_recovery_item_can_be_read() -> None:
    """The upstream reader is not tied to vehicles or electronics domains."""
    from src.upstream import load

    params, case, root = build_everything()
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        assert tables is not None, 'the reader did not recognise the case'

        inflow, composition = tables['inputs'], tables['composition']
        assert set(inflow['Substance_main_parent']) == {PRODUCT}, \
            f"product is {set(inflow['Substance_main_parent'])}, expected {PRODUCT}"
        assert set(composition['Layer 2']) == set(GROUPS), \
            f"groups are {set(composition['Layer 2'])}, expected {set(GROUPS)}"
        assert sorted(inflow['Year']) == list(YEARS), 'both years should be read'
        # Upstream reports kilotonnes whatever the project works in; the
        # conversion to working_unit happens in the engine, on load.
        assert set(inflow['Unit']) == {'kt'}, 'upstream mass is kilotonnes'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_a_resolved_material_lands_at_layer_3() -> None:
    """
    `<element>__<material>__<group>` puts the material at Layer 3 and the
    element beneath it -- where a placeholder used to stand alone.

    Every number here is exact rather than approximate: the fixture writes each
    element as a fixed fraction of its group's mass, so the shares are those
    fractions whatever the draws did.

    The arithmetic being checked, on `Laminate`:

        Si is 0.40 of the group, and 0.30 of the group is Si in `glassfeed`.
        So `glassfeed` is 0.30 of the group and Si is ALL of `glassfeed`;
        the placeholder is the other 0.70, holding the 0.10 of silicon whose
        material upstream does not resolve, and the 0.001 of silver.

    Note what is NOT here: no element or material name appears in `src/`. The
    names in this file are a panel's, and the reader has never seen them.
    """
    from src.upstream import load

    params, case, root = build_everything(materials=True)
    try:
        composition = load(params, params.run.data_folder,
                           quiet=True)['composition']
        one_year = composition[composition['Year'] == YEARS[0]]

        def share(group: str, material: str, element: str = '') -> float:
            rows = one_year[(one_year['Layer 2'] == group)
                            & (one_year['Layer 3'] == material)
                            & (one_year['Layer 4'] == element)]
            assert len(rows) == 1, \
                f'expected one row for {group}/{material}/{element}, got {len(rows)}'
            return float(rows['Value'].iloc[0])

        held = 'Laminate_mixed'
        assert abs(share('Laminate', 'glassfeed') - 0.30) < 1e-6, \
            f"glassfeed is {share('Laminate', 'glassfeed'):.6f} of Laminate, expected 0.30"
        assert abs(share('Laminate', 'glassfeed', 'Si') - 1.0) < 1e-6, \
            'a material holds only the elements exported for it, so they sum to 1'
        assert abs(share('Laminate', held) - 0.70) < 1e-6, \
            f"the placeholder is {share('Laminate', held):.6f}, expected 1 - 0.30"
        assert abs(share('Laminate', held, 'Si') - 0.10 / 0.70) < 1e-6, \
            'the silicon upstream does not resolve should stay in the placeholder'
        assert abs(share('Laminate', held, 'Ag') - 0.001 / 0.70) < 1e-6, \
            'an element with no material resolution is a share of the placeholder'

        # Two materials over one group, one of them holding two elements.
        assert abs(share('JunctionBox', 'braid') - 0.20) < 1e-6
        assert abs(share('JunctionBox', 'solderbed') - 0.065) < 1e-6
        assert abs(share('JunctionBox', 'solderbed', 'Cu') - 0.05 / 0.065) < 1e-6, \
            'within a material, an element is a share of THAT material'

        # Shares still nest: materials within the group, elements within each.
        for (group, _), rows in one_year[one_year['Layer 4'] == ''].groupby(
                ['Layer 2', 'Year']):
            below = rows[rows['Layer 3'] != '']['Value'].sum()
            assert below <= 1.0 + 1e-9, \
                f'{group}: its materials sum to {below:.6f} of it'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_no_material_files_give_exactly_the_rows_they_always_did() -> None:
    """
    The material layer generalises the placeholder rather than replacing it.

    A group upstream does not resolve must come out identical either way, and
    a whole export without the new files must come out as one placeholder at
    1.0 with every element a share of the group -- which is what every case
    read before 2026-08-31, and what the electronics case will still be for any
    domain the new export does not reach.

    `Frame` is in both fixtures and has no material files in either, so it is
    compared row for row across the two.
    """
    from src.upstream import load

    plain_params, plain_case, plain_root = build_everything()
    rich_params, rich_case, rich_root = build_everything(materials=True)
    try:
        plain = load(plain_params, plain_params.run.data_folder,
                     quiet=True)['composition']
        rich = load(rich_params, rich_params.run.data_folder,
                    quiet=True)['composition']

        # Without the files: one placeholder per group, whole, and nothing else
        # at Layer 3.
        materials = plain[(plain['Layer 3'] != '') & (plain['Layer 4'] == '')]
        assert set(materials['Layer 3']) == {f'{g}_mixed' for g in GROUPS}, \
            f"Layer 3 holds {sorted(set(materials['Layer 3']))}, expected placeholders"
        assert (materials['Value'] == 1.0).all(), \
            'with nothing resolved the placeholder must be the whole group'

        columns = ['Year', 'Layer 2', 'Layer 3', 'Layer 4', 'Value']
        for frame in (plain, rich):
            frame.sort_values(columns[:-1], inplace=True)
        untouched = [f[f['Layer 2'] == 'Frame'][columns].reset_index(drop=True)
                     for f in (plain, rich)]
        assert untouched[0].equals(untouched[1]), \
            ('a group with no material files changed when another group gained '
             f'them:\n{untouched[0]}\n{untouched[1]}')
    finally:
        for path in (plain_root, plain_case, rich_root, rich_case):
            shutil.rmtree(path, ignore_errors=True)


def test_parts_exceeding_their_element_are_refused() -> None:
    """
    An element's resolved parts are inside its own total, not beside it. Parts
    claiming more than the whole means the two came from different runs, or
    that the resolution counts something twice -- and either way no share can
    be computed around it.

    Without this the surplus would silently reduce what is left for the
    placeholder, or drive it negative.
    """
    from src.upstream import UpstreamError, load

    params, case, root = build_everything(materials=True)
    try:
        # Si is 0.40 of the Laminate; claim 0.55 of it for one material.
        folder = os.path.join(root, 'upstream', 'HIGH', 'collected')
        whole = np.load(os.path.join(folder, '__domain____Laminate.npy'))
        np.save(os.path.join(folder, 'Si__glassfeed__Laminate.npy'),
                (whole * 0.55).astype(np.float32))

        try:
            load(params, params.run.data_folder, quiet=True)
        except UpstreamError as error:
            message = str(error)
        else:
            raise AssertionError(
                'parts claiming more than their element were read without complaint')

        assert 'Si' in message and 'Laminate' in message, \
            f'the refusal names neither the element nor the group:\n{message}'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_a_case_with_real_materials_solves_and_closes() -> None:
    """
    The point of the layer: a case whose Layer 3 is several real materials and
    a placeholder solves, nests and balances, with coefficients keyed at the
    material layer as readily as at the element layer.

    `tools/make_skeleton.py` writes those coefficients from the composition, so
    the materials reach the TC table without anyone naming them either.
    """
    from src.upstream import load

    params, case, root = build_everything(materials=True)
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        coefficients(case, tables['composition'])

        written = pd.read_csv(os.path.join(case, 'input_data', 'TCs.csv'),
                              dtype=str).fillna('')
        keys = set(written['Input_layer_key'])
        for material in ('glassfeed', 'braid', 'solderbed'):
            assert material in keys, \
                f'{material} reached Layer 3 but no coefficient was written for it'

        solution = RecoveryModelOptimized(
            data_folder=case, layer_names=NAMES, tables=tables,
            working_unit=params.run.working_unit, years='',
        ).solve_models_and_write_to_output()
        solution['Value'] = pd.to_numeric(solution['Value'])

        layers = ['Layer 1', 'Layer 2', 'Layer 3', 'Layer 4']
        depth = (solution[layers] != '').sum(axis=1)
        for year, group in solution.assign(depth=depth).groupby('Year'):
            def total(flow):
                rows = group[group['Stock/Flow ID'] == flow]
                return 0.0 if rows.empty else \
                    rows[rows['depth'] == rows['depth'].min()]['Value'].sum()
            entering = total(FLOW)
            leaving = sum(total(f) for f in ('PV_recovered', 'PV_loss_leaching',
                                             'PV_loss_handling'))
            assert entering > 0, f'{year}: nothing entered'
            assert abs(entering - leaving) / entering < 1e-9, \
                f'{year}: {entering:,.3f} entered but {leaving:,.3f} left'

        # The materials are in the answer, not just in the inputs.
        assert {'glassfeed', 'braid', 'solderbed'} <= set(solution['Layer 3']), \
            'the resolved materials did not survive into the solution'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_a_folder_holding_two_runs_is_refused() -> None:
    """
    An upstream folder is never cleared, so it is the union of every run that
    has written to it: a name a later run does not emit is left behind rather
    than replaced. Read together, one run's element is divided by another run's
    total.

    This is not hypothetical. On 2026-08-31 `element_draws/BAU/collected` held
    four runs at once and Motors' elements came to 1.81 of Motors. It surfaced
    only because `src/rest.py` refuses parts exceeding the whole; a mix landing
    under 1 would have balanced and plotted.

    Nothing in the arrays says which run wrote them. The draw count does: every
    array of one run shares it, and two runs have no reason to.
    """
    from src.upstream import UpstreamError, load

    params, case, root = build_everything()
    try:
        # A leftover under a name the current run does not write, so nothing
        # replaced it -- exactly how the real folder came to hold four runs.
        np.save(os.path.join(root, 'upstream', 'HIGH', 'collected',
                             f'Zz__{GROUPS[0]}.npy'),
                np.full((DRAWS // 4, len(YEARS)), 0.1, dtype=np.float32))

        try:
            load(params, params.run.data_folder, quiet=True)
        except UpstreamError as error:
            message = str(error)
        else:
            raise AssertionError(
                'a folder holding two draw counts was read without complaint')

        for wanted in (f'{DRAWS:,}', f'{DRAWS // 4:,}', f'Zz__{GROUPS[0]}'):
            assert wanted in message, \
                f'the refusal does not mention {wanted!r}:\n{message}'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_products_from_different_runs_are_refused() -> None:
    """
    The same mix one level up. 04_01's five drivetrains are five folders, so
    re-running one alone leaves the other four at the previous width -- and
    each folder is internally consistent, which is what the check inside
    `read_draws` would see.
    """
    from src.upstream import UpstreamError, load

    params, case, root = build_everything(products=PRODUCTS)
    try:
        narrowed = os.path.join(root, 'upstream', 'HIGH', f'{PRODUCTS[0]}_collected')
        for name in sorted(os.listdir(narrowed)):
            path = os.path.join(narrowed, name)
            np.save(path, np.load(path)[:DRAWS // 4])

        try:
            load(params, params.run.data_folder, quiet=True)
        except UpstreamError as error:
            message = str(error)
        else:
            raise AssertionError(
                'products at two draw counts were read without complaint')

        assert PRODUCTS[0] in message and PRODUCTS[1] in message, \
            f'the refusal does not name the products that disagree:\n{message}'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_a_different_recovery_item_solves_and_closes() -> None:
    """
    The whole pipeline, on a non-vehicle item: read, derive rest, solve, and
    check that what enters equals what leaves in every year.
    """
    from src.upstream import load

    params, case, root = build_everything()
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        coefficients(case, tables['composition'])

        # Two of the three groups are incomplete, and a rest is derived per
        # parent per year -- the same rule the vehicle case relies on.
        with_rest, notes = add_rest(tables['composition'])
        assert len(notes) == 2 * len(YEARS), \
            f'expected 2 incomplete groups x {len(YEARS)} years, got {len(notes)}'

        solution = RecoveryModelOptimized(
            data_folder=case, layer_names=NAMES, tables=tables,
            working_unit=params.run.working_unit, years='',
        ).solve_models_and_write_to_output()
        solution['Value'] = pd.to_numeric(solution['Value'])

        layers = ['Layer 1', 'Layer 2', 'Layer 3', 'Layer 4']
        depth = (solution[layers] != '').sum(axis=1)
        for year, group in solution.assign(depth=depth).groupby('Year'):
            def total(flow):
                rows = group[group['Stock/Flow ID'] == flow]
                return 0.0 if rows.empty else \
                    rows[rows['depth'] == rows['depth'].min()]['Value'].sum()
            entering = total(FLOW)
            leaving = sum(total(f) for f in ('PV_recovered', 'PV_loss_leaching',
                                             'PV_loss_handling'))
            assert entering > 0, f'{year}: nothing entered'
            assert abs(entering - leaving) / entering < 1e-9, \
                f'{year}: {entering:g} in, {leaving:g} out'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_children_can_be_materials_instead_of_elements() -> None:
    """
    The 04_01 shape: the upstream child is a material, not an element.

    Same fixture, same files, one line of source.csv different. The children
    must land at Layer 3 with Layer 4 empty and no placeholder material
    invented -- and the case must still solve and close.
    """
    from src.upstream import load

    params, case, root = build_everything(child_layer='material')
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        composition = tables['composition']

        assert set(composition['Layer 4']) == {''}, \
            f"Layer 4 should be unused, found {sorted(set(composition['Layer 4']))}"

        children = {e for group in ELEMENTS.values() for e in group}
        at_three = set(composition['Layer 3']) - {''}
        assert at_three == children, \
            f'Layer 3 holds {sorted(at_three)}, expected {sorted(children)}'
        assert not any(name.endswith('_mixed') for name in at_three), \
            'a placeholder material was invented where the child is already one'

        coefficients(case, composition)
        solution = RecoveryModelOptimized(
            data_folder=case, layer_names=NAMES, tables=tables,
            working_unit=params.run.working_unit, years='',
        ).solve_models_and_write_to_output()
        solution['Value'] = pd.to_numeric(solution['Value'])

        layers = ['Layer 1', 'Layer 2', 'Layer 3', 'Layer 4']
        depth = (solution[layers] != '').sum(axis=1)
        for year, group in solution.assign(depth=depth).groupby('Year'):
            def total(flow):
                rows = group[group['Stock/Flow ID'] == flow]
                return 0.0 if rows.empty else \
                    rows[rows['depth'] == rows['depth'].min()]['Value'].sum()
            entering = total(FLOW)
            leaving = sum(total(f) for f in ('PV_recovered', 'PV_loss_leaching',
                                             'PV_loss_handling'))
            assert entering > 0, f'{year}: nothing entered'
            assert abs(entering - leaving) / entering < 1e-9, \
                f'{year}: {entering:g} in, {leaving:g} out'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_several_products_close_on_their_own_totals() -> None:
    """
    The 04_01 drivetrain shape: one case, several Layer 1 values.

    Each product is exported to its own folder at its own level, and each must
    come out as its own whole. Pooling them would still balance and still plot
    -- every share would just be wrong by the ratio of one product's mass to
    all of them -- so the check is on the shares, not on the total.
    """
    from src.upstream import load

    params, case, root = build_everything(products=PRODUCTS)
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        inflow, composition = tables['inputs'], tables['composition']

        assert set(inflow['Substance_main_parent']) == set(PRODUCTS), \
            f"products are {set(inflow['Substance_main_parent'])}, expected {PRODUCTS}"
        assert len(inflow) == len(PRODUCTS) * len(YEARS), \
            f'expected one inflow row per product per year, got {len(inflow)}'

        # The folders were written at 40 kt and 80 kt, so the products must
        # differ -- identical totals would mean one folder was read twice.
        totals = inflow.groupby('Substance_main_parent')['Value'].sum()
        assert totals.max() / totals.min() > 1.5, \
            f'the products came out the same size ({dict(totals)}); one folder read twice?'

        # The component share is a share of its OWN product.
        top = composition[composition['Layer 3'] == '']
        for (product, year), rows in top.groupby(['Layer 1', 'Year']):
            assert abs(rows['Value'].sum() - 1.0) < 1e-6, \
                f'{product} {year}: component shares sum to {rows["Value"].sum():.6f}'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_a_workbook_case_solves_the_same_as_a_csv_case() -> None:
    """
    The three tables may be a case.xlsx with three sheets instead of three
    CSVs, and the answer must not move.

    Excel is offered because the coefficient table is filled in by hand and a
    spreadsheet is a better surface for that -- dropdowns instead of typed
    flow names, a note beside a number. None of that is worth a different
    result, so this solves the same fixture both ways and compares.

    It also pins the type behaviour, which is where this went wrong first. A
    CSV read with keep_default_na=False gives '' and '1' as strings; Excel has
    no empty string, so the column arrives float64 as nan and 1.0, and
    `is_residual` -- tested with str(value) in ('1', 'True', 'true') -- stops
    matching anything. The Monte Carlo then refuses a group whose residual rows
    it cannot identify. src/case_tables.normalise exists for exactly this.
    """
    from src import case_tables
    from src.upstream import load

    params, case, root = build_everything()
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        coefficients(case, tables['composition'])

        def solve():
            return RecoveryModelOptimized(
                data_folder=case, layer_names=NAMES, tables=tables,
                working_unit=params.run.working_unit, years='',
            ).solve_models_and_write_to_output()

        from_csv = solve()

        # Same table, moved into a workbook sheet; the CSV goes, so nothing
        # can silently read the old one.
        written = case_tables.read(case, 'TCs')
        case_tables.write_sheet(case, 'TCs', written)
        os.unlink(case_tables.csv_path(case, 'TCs'))
        assert case_tables.where(case, 'TCs')[0] == 'xlsx'

        from_workbook = solve()

        assert len(from_csv) == len(from_workbook), \
            f'{len(from_csv)} rows from CSV, {len(from_workbook)} from the workbook'

        keys = ['Year', 'Stock/Flow ID', 'Layer 1', 'Layer 2', 'Layer 3', 'Layer 4']
        merged = from_csv.merge(from_workbook, on=keys, how='outer',
                                suffixes=('_csv', '_xlsx'), indicator=True)
        assert (merged['_merge'] == 'both').all(), \
            'the two runs produced different rows'

        worst = (pd.to_numeric(merged['Value_csv'])
                 - pd.to_numeric(merged['Value_xlsx'])).abs().max()
        assert worst < 1e-9, f'workbook and CSV differ by {worst:g}'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)



def test_a_dead_layer_column_is_not_written() -> None:
    """
    A case writes only the layers it actually reaches.

    All three real cases are material-keyed, so Layer 4 is empty in every row of
    every one of them, and each was writing a dead column into its solution, its
    summary and two workbook sheets. Dropped at the moment of WRITING only.

    Both halves matter, so both are checked here:
      - the file loses the column when nothing fills it;
      - the FRAME returned keeps it, because 03_run_monte_carlo.py merges the
        deterministic answer onto the Monte Carlo one using all four layers.
        Dropping before returning would break that join rather than tidy a file.
    """
    from src.upstream import load

    for child_layer, expected in (('material', False), ('element', True)):
        params, case, root = build_everything(child_layer=child_layer)
        try:
            tables = load(params, params.run.data_folder, quiet=True)
            coefficients(case, tables['composition'])
            solution = RecoveryModelOptimized(
                data_folder=case, layer_names=NAMES, tables=tables,
                working_unit=params.run.working_unit, years='',
            ).solve_models_and_write_to_output()

            assert 'Layer 4' in solution.columns, (
                f'{child_layer}: the returned frame lost Layer 4, which the '
                f'Monte Carlo merges on')

            written = pd.read_csv(f'{case}/output_data/solution_optimized_model.csv',
                                  keep_default_na=False, na_values=[])
            has = 'Layer 4' in written.columns
            assert has == expected, (
                f"{child_layer}: solution_optimized_model.csv "
                f"{'has' if has else 'lacks'} a Layer 4 column, expected "
                f"{'one' if expected else 'none'}")
            if expected:
                assert (written['Layer 4'] != '').any(), (
                    'Layer 4 was written but is empty in every row')
        finally:
            shutil.rmtree(root, ignore_errors=True)
            shutil.rmtree(case, ignore_errors=True)


def test_a_layer_with_a_gap_below_it_is_kept() -> None:
    """
    Only TRAILING layers are dropped.

    An empty layer with a filled one beneath it is a gap in the nesting, which
    `validate_inputs` refuses at input. If one ever reached the writer, dropping
    it would quietly restate the nesting as something it is not -- so the rule
    is "keep everything up to the deepest filled layer", not "drop every empty
    column".
    """
    from src.rest import drop_unused_layers

    gap = pd.DataFrame({'Layer 1': ['P'], 'Layer 2': ['C'], 'Layer 3': [''],
                        'Layer 4': ['E'], 'Value': [1.0]})
    assert list(drop_unused_layers(gap).columns) == list(gap.columns), \
        'a gap in the nesting was silently closed up'


def test_the_monte_carlo_runs_on_a_different_item() -> None:
    """Sampling, the sum-to-1 groups and the spread all work off vehicle data."""
    from src.upstream import load

    params, case, root = build_everything()
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        coefficients(case, tables['composition'])

        run = solve_draws(case, NAMES, draws=200, seed=0, tables=tables, years='')
        assert run.report['uncertain'], 'the ranges did not reach the sampler'
        assert run.values.std(axis=1).max() > 0, 'the Monte Carlo produced no spread'
        assert run.report['groups'] > 0, 'no constrained groups were found'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_the_source_sheet_offers_the_child_layer_choices() -> None:
    """
    Writing `source` constrains the cell beside `child_layer` to the two
    layers, wherever in the sheet that row happens to sit -- and writing the
    sheet again does not accumulate a second copy of every list.

    child_layer is refused on load if it is wrong, so this buys no safety. It
    buys the sheet SAYING what the choices are, which matters because the two
    are not guessable and the wrong one still balances.
    """
    import openpyxl

    from src import case_tables
    from src.source import CHILD_LAYERS

    case = tempfile.mkdtemp(prefix='source-dropdown-')
    try:
        # child_layer LAST, so a hard-coded row number cannot pass by luck.
        frame = pd.DataFrame([
            {'key': 'product', 'value': 'BEV'},
            {'key': 'flow', 'value': 'collected'},
            {'key': 'child_layer', 'value': 'material'},
        ])
        case_tables.write_sheet(case, 'source', frame)

        book = openpyxl.load_workbook(case_tables.workbook_path(case))
        rules = book['source'].data_validations.dataValidation
        assert len(rules) == 1, f'{len(rules)} validations, expected 1'

        cells = [str(area) for rule in rules for area in rule.sqref.ranges]
        assert cells == ['B4'], f'constrained {cells}, expected the B4 cell'

        # The rule points at a range on the hidden sheet, never an inline list:
        # Excel truncates an inline list at 255 characters without saying so.
        assert case_tables.LISTS_SHEET in rules[0].formula1
        assert book[case_tables.LISTS_SHEET].sheet_state == 'hidden'

        header = [c.value for c in book[case_tables.LISTS_SHEET][1]]
        column = header.index('child_layer') + 1
        offered = tuple(
            book[case_tables.LISTS_SHEET].cell(row=line, column=column).value
            for line in (2, 3))
        assert offered == CHILD_LAYERS, f'offers {offered}, not {CHILD_LAYERS}'

        # Writing again must reuse the column rather than append another.
        width = book[case_tables.LISTS_SHEET].max_column
        for _ in range(3):
            case_tables.write_sheet(case, 'source',
                                    case_tables.read(case, 'source'))
        again = openpyxl.load_workbook(case_tables.workbook_path(case))
        grown = again[case_tables.LISTS_SHEET].max_column
        assert grown == width, \
            f'_lists grew from {width} to {grown} columns over three rewrites'
        assert len(again['source'].data_validations.dataValidation) == 1

        # And a sheet that never mentions child_layer gets no rule at all.
        plain = pd.DataFrame([{'key': 'product', 'value': 'BEV'}])
        case_tables.write_sheet(case, 'source', plain)
        bare = openpyxl.load_workbook(case_tables.workbook_path(case))
        assert not bare['source'].data_validations.dataValidation
    finally:
        shutil.rmtree(case, ignore_errors=True)


def test_every_written_sheet_marks_its_header_row() -> None:
    """Row 1 comes out bold on a fill, so it reads as headings, not as data."""
    import openpyxl

    from src import case_tables

    case = tempfile.mkdtemp(prefix='header-style-')
    try:
        frame = pd.DataFrame([{'key': 'product', 'value': 'BEV'}])
        case_tables.write_sheet(case, 'source', frame)
        case_tables.write_sheet(case, 'processes', pd.DataFrame(
            [{'Input_FlowID': 'a', 'Output_FlowID': 'b'}]))

        book = openpyxl.load_workbook(case_tables.workbook_path(case))
        for name in ('source', 'processes'):
            sheet = book[name]
            for cell in sheet[1]:
                assert cell.font.bold, f'{name}!{cell.coordinate} is not bold'
                assert cell.fill.start_color.rgb.endswith(case_tables.HEADER_FILL), \
                    f'{name}!{cell.coordinate} fill is {cell.fill.start_color.rgb}'
            # The first data row must NOT be styled, or nothing stands out.
            assert not sheet['A2'].font.bold, f'{name}: row 2 is bold too'
    finally:
        shutil.rmtree(case, ignore_errors=True)


def test_role_alone_decides_what_a_flow_counts_as() -> None:
    """
    `role` is the only column consulted. There is no `is_loss` to agree or
    disagree with it, and a role that is missing or misspelled is refused by
    name rather than quietly counted as recovered -- which is the failure the
    column was added to prevent in the first place.
    """
    from src.rest import flow_roles

    case = tempfile.mkdtemp(prefix='roles-')
    os.makedirs(os.path.join(case, 'input_data'), exist_ok=True)
    rows = [
        dict(Input_FlowID='a', Output_FlowID='b', process='p', technology='t',
             keyed_at='component', role='intermediate'),
        dict(Input_FlowID='a', Output_FlowID='a_loss', process='p', technology='t',
             keyed_at='component', role='loss'),
        dict(Input_FlowID='b', Output_FlowID='c', process='q', technology='t',
             keyed_at='element', role='recovered'),
    ]
    try:
        path = os.path.join(case, 'input_data', 'processes.csv')
        pd.DataFrame(rows).to_csv(path, index=False)
        roles = flow_roles(case)
        assert roles == {'b': 'intermediate', 'a_loss': 'loss', 'c': 'recovered'}, roles

        # A misspelling must not silently become 'recovered'.
        bad = [dict(row) for row in rows]
        bad[2]['role'] = 'recoverd'
        pd.DataFrame(bad).to_csv(path, index=False)
        try:
            flow_roles(case)
        except ValueError as error:
            assert 'recoverd' in str(error), error
            assert 'c' in str(error), error
        else:
            raise AssertionError("a misspelled role was accepted")

        # So must a blank one.
        blank = [dict(row) for row in rows]
        blank[2]['role'] = ''
        pd.DataFrame(blank).to_csv(path, index=False)
        try:
            flow_roles(case)
        except ValueError:
            pass
        else:
            raise AssertionError('a blank role was accepted')
    finally:
        shutil.rmtree(case, ignore_errors=True)


def test_the_skeleton_counts_loss_destinations_from_role() -> None:
    """
    make_skeleton fills a `rest` row in only where a flow has exactly one loss
    destination, and it now learns that from `role`, not from `is_loss`.
    With two losses the split is a judgement it will not make.
    """
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
    from tools.make_skeleton import loss_destinations_of

    one = pd.DataFrame([
        dict(Input_FlowID='a', Output_FlowID='b', role='intermediate'),
        dict(Input_FlowID='a', Output_FlowID='a_loss', role='loss'),
    ])
    assert loss_destinations_of(one) == {'a': 1}, loss_destinations_of(one)

    two = pd.DataFrame([
        dict(Input_FlowID='a', Output_FlowID='a_loss1', role='loss'),
        dict(Input_FlowID='a', Output_FlowID='a_loss2', role='loss'),
    ])
    assert loss_destinations_of(two) == {'a': 2}, loss_destinations_of(two)

    # A handoff is not a loss: material leaves, but to another model.
    handed = pd.DataFrame([
        dict(Input_FlowID='a', Output_FlowID='elsewhere', role='handoff'),
    ])
    assert loss_destinations_of(handed) == {'a': 0}, loss_destinations_of(handed)


def test_a_resource_that_cannot_leave_its_flow_is_refused() -> None:
    """
    A resource reaching a non-terminal flow with no coefficient to leave by
    loses its mass, silently: the run still writes a solution, still draws its
    figures, and reports a recovery rate computed from less mass than entered.

    Measured 2026-09-01 on the electronics case at 5.9%, and found by totalling
    the terminal flows by hand rather than by anything complaining. Four
    resources had been moved to a new material by the Layer 3 work, so the
    coefficients keyed on their old one no longer reached them.

    The opposite direction -- a coefficient naming a resource that does not
    exist -- has been a warning since 2026-08-17, and correctly: an inert row
    costs nothing. Same join read the other way, and it costs the answer.

    This lives here rather than beside the other validation tests because the
    check reads which flows are terminal from the `processes` table, and no
    reference fixture has one. This suite's synthetic case does.
    """
    from src.upstream import load
    from src.validate_inputs import InputDataError, validate

    params, case, root = build_everything()
    try:
        tables = load(params, params.run.data_folder, quiet=True)
        coefficients(case, tables['composition'])

        path = os.path.join(case, 'input_data', 'TCs.csv')
        tcs = pd.read_csv(path, keep_default_na=False, na_values=[])
        # Take away every way out of the delaminated flow for one element,
        # leaving the composition untouched. Its mass now arrives and stays.
        stuck = ((tcs['Input_FlowID'] == 'PV_delaminated')
                 & (tcs['TC_target_key'] == 'Cu'))
        assert stuck.any(), 'the fixture no longer has the rows this test removes'
        tcs[~stuck].to_csv(path, index=False)

        try:
            # The panel's inflow and composition are never written to disk, so
            # the tables are handed in the way a stage hands them in.
            validate(case, tables)
        except InputDataError as error:
            message = str(error)
        else:
            raise AssertionError(
                'a resource with no way out of a non-terminal flow was accepted')

        assert 'Cu' in message and 'PV_delaminated' in message, \
            f'the refusal names neither the resource nor the flow:\n{message}'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_the_skeleton_never_deletes_a_filled_row() -> None:
    """
    A re-run must not remove a coefficient somebody entered, whatever happened
    to the composition since.

    This is not hypothetical tidiness. A resource leaves the composition for
    two ordinary reasons, and neither says the row is wrong: narrowing `groups`
    to work one component at a time -- the workflow the script exists for --
    and an upstream export resolving fewer elements than the last one. Both
    happened together on 2026-09-01 and a single re-run deleted 32 filled rows
    from `bev_electronics`, every rare earth among them, reporting it as
    `dropped`.

    A blank stale row is still removed: it says nothing, so there is nothing to
    lose.
    """
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
    from tools.make_skeleton import merge

    columns = ['Input_FlowID', 'Input_layer', 'Input_layer_key', 'Output_FlowID',
               'TC_target_layer', 'TC_target_key', 'value', 'value_min',
               'value_max', 'is_residual', 'process', 'technology', 'source']

    def row(key, value, **rest):
        base = dict.fromkeys(columns, '')
        base.update(Input_FlowID='F_in', Output_FlowID='F_out',
                    Input_layer_key='old_material', TC_target_key=key,
                    value=value, **rest)
        return base

    existing = pd.DataFrame([
        row('Keeper', '0.42', source='Smith 2023, table 4'),   # filled, now absent
        row('Blank', ''),                                      # blank, now absent
        row('Present', '0.30'),                                # filled and still there
    ], columns=columns)
    skeleton = pd.DataFrame([
        row('Present', ''),
        row('Fresh', ''),
    ], columns=columns)

    merged, change = merge(existing, skeleton)
    keys = list(merged['TC_target_key'])

    assert 'Keeper' in keys, \
        f'a filled row was deleted because its resource left the composition: {keys}'
    assert 'Blank' not in keys, \
        'a blank stale row carries nothing and should not be kept'
    assert keys[-1] == 'Keeper', \
        f'the inert row should sit at the end, out of the way: {keys}'
    assert change['inert'] == 1, f"expected one inert row, got {change['inert']}"

    kept = merged[merged['TC_target_key'] == 'Keeper'].iloc[0]
    assert kept['value'] == '0.42' and kept['source'] == 'Smith 2023, table 4', \
        'the inert row was kept but its value or its provenance was not'

    # And the ordinary business still works.
    assert merged[merged['TC_target_key'] == 'Present'].iloc[0]['value'] == '0.30', \
        'a value that is still in the composition was overwritten'
    assert change['kept'] == 1 and change['added'] == 1, change


def test_the_processes_sheet_offers_role_and_keyed_at() -> None:
    """
    Both fixed-vocabulary columns come out as dropdowns over the data rows,
    and the header row is left alone. `role` matters most: it has no default,
    so a misspelling stops the run -- and a list you pick from cannot be
    misspelled.
    """
    import openpyxl

    from src import case_tables
    from src.rest import KEYED_AT, ROLES

    case = tempfile.mkdtemp(prefix='processes-dropdown-')
    try:
        frame = pd.DataFrame([
            dict(Input_FlowID='a', Output_FlowID='b', process='p',
                 technology='t', keyed_at='component', role='intermediate'),
            dict(Input_FlowID='a', Output_FlowID='a_loss', process='p',
                 technology='t', keyed_at='component', role='loss'),
        ])
        case_tables.write_sheet(case, 'processes', frame)

        book = openpyxl.load_workbook(case_tables.workbook_path(case))
        sheet = book['processes']
        ranges = {}
        for rule in sheet.data_validations.dataValidation:
            for area in rule.sqref.ranges:
                ranges[str(area)] = rule.formula1

        # role is column F, keyed_at column E; rows 2-3 are the data.
        assert set(ranges) == {'E2:E3', 'F2:F3'}, sorted(ranges)

        lists = book[case_tables.LISTS_SHEET]
        header = [cell.value for cell in lists[1]]
        for name, expected in (('role', ROLES), ('keyed_at', KEYED_AT)):
            column = header.index(name) + 1
            offered = tuple(lists.cell(row=line, column=column).value
                            for line in range(2, len(expected) + 2))
            assert offered == tuple(expected), f'{name} offers {offered}'
    finally:
        shutil.rmtree(case, ignore_errors=True)


def test_the_keyed_at_vocabulary_has_one_definition() -> None:
    """
    make_skeleton derives each layer's parent from the nesting, so the layers
    it accepts are the layers src.rest declares. Two lists that must agree are
    two lists that will not -- the point of deriving one from the other.
    """
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
    from tools.make_skeleton import INPUT_LAYER_FOR, LAYER_COLUMN
    from src.rest import KEYED_AT

    assert tuple(INPUT_LAYER_FOR) == KEYED_AT, \
        f'make_skeleton accepts {tuple(INPUT_LAYER_FOR)}, rest declares {KEYED_AT}'
    # Every parent is the layer directly above its child.
    order = list(LAYER_COLUMN)
    for child, parent in INPUT_LAYER_FOR.items():
        assert order.index(parent) == order.index(child) - 1, \
            f'{child} reads from {parent}, which is not the layer above it'


# ----------------------------------------------------------------------
#  The Sankeys must say which year they show, and in the right unit
# ----------------------------------------------------------------------

def test_the_sankey_names_the_combination_it_shows() -> None:
    """
    `replay` took `input_data[0]` and said nothing. The electronics case has
    five years, so every Sankey in figures/ described 2030 while every other
    output was headlined 2050, and nothing on the figure said which it was.

    DEFECTS.md called this low severity because the fixtures had one
    combination each. The real case grew to five and it started biting
    unnoticed, which is the argument for the figure naming its own subject
    rather than for choosing a better default.
    """
    from src import plot_flows

    params, case, root = build_everything()
    try:
        from src.upstream import load
        tables = load(params, case, quiet=True)
        coefficients(case, tables['composition'])
        chosen, entries = plot_flows.chosen_entry(case, tables)
        assert entries >= 1, 'no combinations at all'
        assert chosen, 'the chosen combination has no label'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_the_sankey_labels_the_unit_it_actually_drew() -> None:
    """
    The unit came from the inputs table's own `Unit` column -- the SOURCE unit,
    kt -- while the values had been converted to `run.working_unit`, kg. Every
    Sankey was therefore labelled a factor of a million out: aluminium in the
    2030 electronics case printed as `887,760.1 kt` when it is 887,760 kg.

    A wrong unit is the one mistake this project treats as serious, because it
    is invisible: the number looks fine and the reader supplies the meaning.
    """
    from src import plot_flows

    params, case, root = build_everything()
    try:
        from src.upstream import load
        tables = load(params, case, quiet=True)
        source_unit = tables['inputs']['Unit'].iloc[0]
        drawn = plot_flows.unit_drawn(params)
        assert drawn == params.run.working_unit, \
            f'labelled {drawn!r}, drew {params.run.working_unit!r}'
        if source_unit != params.run.working_unit:
            assert drawn != source_unit, \
                f'still labelling with the source unit {source_unit!r}'
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(case, ignore_errors=True)


def test_the_worklist_agrees_with_what_the_run_refuses() -> None:
    """
    tc_worklist said "REFUSED at run time" for any group holding a fixed row,
    while sampling refuses only a group with exactly ONE row free. A case with
    a legitimate fixed row -- a route that does not apply, written as 0 -- was
    reported as broken and then ran perfectly, which is the worst way for a
    warning to be wrong: it teaches you to ignore it.

    The two must say the same thing, so this asks both.
    """
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
    import numpy as np

    from src.sampling import SamplingError, numeric_bounds, sample
    from tools.tc_worklist import COLLAPSED, classify

    def group(*bands) -> pd.DataFrame:
        common = dict(Input_FlowID='F_in', Input_layer='Layer 3',
                      Input_layer_key='M', TC_target_layer='Layer 4',
                      TC_target_key='E')
        return pd.DataFrame([
            dict(**common, Output_FlowID=f'F_{n}', value=str(mode),
                 value_min=str(lo), value_max=str(hi), is_residual='')
            for n, (lo, mode, hi) in enumerate(bands)])

    shapes = {
        'one free row, one fixed': ((0.02, 0.10, 0.30), (0.90, 0.90, 0.90)),
        'two free rows, one fixed': ((0.30, 0.50, 0.70), (0.20, 0.35, 0.50),
                                     (0.15, 0.15, 0.15)),
        'all free': ((0.30, 0.50, 0.70), (0.25, 0.35, 0.50), (0.05, 0.15, 0.30)),
        'all fixed': ((0.6, 0.6, 0.6), (0.4, 0.4, 0.4)),
    }

    for name, bands in shapes.items():
        tcs = numeric_bounds(group(*bands))
        low = tcs['value_min'].to_numpy(float)
        mode = tcs['value'].to_numpy(float)
        high = tcs['value_max'].to_numpy(float)
        members = np.arange(len(tcs))
        status, _ = classify(low, mode, high, None, members)

        try:
            sample(group(*bands), draws=128, rule='condition')
            refused = False
        except SamplingError:
            refused = True

        assert (status == COLLAPSED) == refused, (
            f'{name}: the worklist says {status!r} but the run '
            f'{"refuses" if refused else "accepts"} it')


def test_the_structure_diagram_states_each_endpoints_role():
    """
    An endpoint's ROLE reaches the diagram, and `intermediate` does not.

    Two boxes with no arrow leaving them are indistinguishable on the picture
    -- one may be gold coming back, the other a shredded board this case does
    not follow -- and until 2026-09-02 the diagram said only which layer each
    was expressed at. `intermediate` is left off on purpose: it says an arrow
    leaves the box, which the arrow already says.
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    from src.plot_structure import ROLE_LABEL, render

    tcs = pd.DataFrame([
        dict(Input_FlowID='F_in', Input_layer='product', Input_layer_key='P',
             Output_FlowID='F_middle', TC_target_layer='component',
             TC_target_key='C', value=1.0, value_min=1.0, value_max=1.0,
             process='taking_apart', technology='manual'),
        dict(Input_FlowID='F_middle', Input_layer='component', Input_layer_key='C',
             Output_FlowID='F_back', TC_target_layer='material',
             TC_target_key='M', value=0.9, value_min=0.8, value_max=1.0,
             process='refining', technology='chemical'),
        dict(Input_FlowID='F_middle', Input_layer='component', Input_layer_key='C',
             Output_FlowID='F_gone', TC_target_layer='material',
             TC_target_key='M', value=0.1, value_min=0.0, value_max=0.2,
             process='refining', technology='chemical'),
        dict(Input_FlowID='F_middle', Input_layer='component', Input_layer_key='C',
             Output_FlowID='F_elsewhere', TC_target_layer='material',
             TC_target_key='M', value=1.0, value_min=1.0, value_max=1.0,
             process='shipping', technology='definitional'),
    ])
    roles = {'F_middle': 'intermediate', 'F_back': 'recovered',
             'F_gone': 'loss', 'F_elsewhere': 'handoff'}

    figure = render(tcs, 'a_case', roles=roles)
    drawn = {text.get_text() for text in figure.axes[0].texts}
    plt.close(figure)

    for flow, role in roles.items():
        if role == 'intermediate':
            continue
        assert ROLE_LABEL[role] in drawn, (
            f'{flow} is {role} and the diagram does not say so')
    assert 'intermediate' not in ' '.join(drawn), (
        'intermediate is on the diagram; the arrow already says it')

    # And a case with no processes table keeps the diagram it had.
    plain = render(tcs, 'a_case')
    without = {text.get_text() for text in plain.axes[0].texts}
    plt.close(plain)
    assert not ({ROLE_LABEL[r] for r in ROLE_LABEL} & without), (
        'a case with no roles got role lines anyway')


def main() -> int:
    tests = [value for name, value in sorted(globals().items())
             if name.startswith('test_') and callable(value)]
    failures = 0
    for test in tests:
        try:
            test()
        except AssertionError as error:
            failures += 1
            print(f'FAIL  {test.__name__}\n      {error}\n')
        except Exception:
            failures += 1
            print(f'ERROR {test.__name__}')
            traceback.print_exc()
        else:
            print(f'ok    {test.__name__}')

    print(f'\n{len(tests) - failures} of {len(tests)} passed')
    return 1 if failures else 0


if __name__ == '__main__':
    raise SystemExit(main())
